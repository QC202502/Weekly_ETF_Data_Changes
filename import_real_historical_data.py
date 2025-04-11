#!/usr/bin/env python3
"""
导入真实ETF历史数据

此脚本从Excel文件中读取所有真实ETF自选人数和持有人数据，并导入数据库
"""

import os
import pandas as pd
import xlrd
import re
import struct
import binascii
import openpyxl
from datetime import datetime
from database.models import Database
import subprocess
import sys

# 尝试导入pyexcel包，如果不存在则安装
try:
    import pyexcel
except ImportError:
    print("正在安装pyexcel...")
    subprocess.check_call([sys.executable, "-m", "pip", "install", "pyexcel"])
    import pyexcel

# 尝试导入pyexcel-xlsx包，如果不存在则安装
try:
    import pyexcel_xlsx
except ImportError:
    print("正在安装pyexcel-xlsx...")
    subprocess.check_call([sys.executable, "-m", "pip", "install", "pyexcel-xlsx"])
    import pyexcel_xlsx

# 尝试导入pyexcel-xls包，如果不存在则安装
try:
    import pyexcel_xls
except ImportError:
    print("正在安装pyexcel-xls...")
    subprocess.check_call([sys.executable, "-m", "pip", "install", "pyexcel-xls"])
    import pyexcel_xls

def extract_date_from_filename(filename):
    """从文件名中提取日期"""
    # 尝试查找格式为YYYYMMDD的日期
    date_match = re.search(r'(\d{8})', filename)
    if date_match:
        date_str = date_match.group(1)
        try:
            # 将YYYYMMDD转换为YYYY-MM-DD
            date_obj = datetime.strptime(date_str, '%Y%m%d')
            return date_obj.strftime('%Y-%m-%d')
        except ValueError:
            pass
    return None

def read_binary_file_head(file_path, size=100):
    """读取文件头部的二进制数据，检查文件格式"""
    try:
        with open(file_path, 'rb') as f:
            binary_data = f.read(size)
            hex_data = binascii.hexlify(binary_data).decode('utf-8')
            print(f"文件头部二进制数据: {hex_data}")
            
            # 检查Excel签名
            # XLSX (Office 2007+) 签名: 50 4B 03 04 (PK..)
            if binary_data.startswith(b'PK\x03\x04'):
                print("文件签名显示这是一个XLSX文件(Office 2007+)")
                return "xlsx"
            # XLS (Office 97-2003) 签名: D0 CF 11 E0 A1 B1 1A E1
            elif binary_data.startswith(b'\xD0\xCF\x11\xE0\xA1\xB1\x1A\xE1'):
                print("文件签名显示这是一个XLS文件(Office 97-2003)")
                return "xls"
            else:
                print(f"未知文件格式")
                return "unknown"
    except Exception as e:
        print(f"读取文件二进制数据时出错: {str(e)}")
        return "error"

def read_with_pyexcel(file_path):
    """使用pyexcel读取Excel文件"""
    print(f"使用pyexcel读取: {file_path}")
    try:
        # 使用pyexcel读取Excel文件
        data = pyexcel.get_array(file_name=file_path)
        
        if len(data) > 0:
            print(f"总行数: {len(data)}")
            print(f"第一行: {data[0]}")
            
            # 获取列名（第一行）
            headers = data[0] if data else []
            
            # 创建数据列表（跳过第一行）
            rows = []
            for row in data[1:]:
                # 创建行数据字典
                row_data = {}
                for i, value in enumerate(row):
                    if i < len(headers):
                        row_data[headers[i]] = value
                    else:
                        row_data[f"Column_{i}"] = value
                rows.append(row_data)
            
            # 创建DataFrame
            df = pd.DataFrame(rows)
            
            print(f"读取到 {len(df)} 行数据")
            return df
        else:
            print("文件内容为空")
            return None
    except Exception as e:
        print(f"使用pyexcel读取Excel文件时出错: {str(e)}")
        return None

def read_with_raw_openpyxl(file_path):
    """使用原生openpyxl读取XLSX文件，直接访问工作表单元格"""
    print(f"使用原生openpyxl读取文件: {file_path}")
    try:
        # 打开Excel文件
        workbook = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        
        # 获取第一个工作表
        if len(workbook.sheetnames) > 0:
            sheet = workbook[workbook.sheetnames[0]]
            print(f"工作表名称: {sheet.title}")
            
            # 确定有效行数和列数
            max_row = 0
            max_col = 0
            for row_idx, row in enumerate(sheet.iter_rows()):
                if row_idx == 0:  # 获取最大列数
                    for cell in row:
                        if cell.value is not None:
                            max_col = max(max_col, cell.column)
                if any(cell.value is not None for cell in row):
                    max_row = max(max_row, row_idx + 1)
            
            print(f"检测到有效行数: {max_row}, 有效列数: {max_col}")
            
            if max_row == 0 or max_col == 0:
                print("没有找到有效数据")
                return None
            
            # 获取列名
            headers = []
            for col_idx in range(1, max_col + 1):
                cell_value = sheet.cell(row=1, column=col_idx).value
                headers.append(cell_value if cell_value is not None else f"Column_{col_idx}")
            
            print(f"列名: {headers}")
            
            # 读取数据
            data = []
            for row_idx in range(2, max_row + 1):  # 从第2行开始（跳过表头）
                row_data = {}
                all_none = True
                for col_idx in range(1, max_col + 1):
                    cell_value = sheet.cell(row=row_idx, column=col_idx).value
                    if cell_value is not None:
                        all_none = False
                    row_data[headers[col_idx - 1]] = cell_value
                
                if not all_none:  # 只添加至少有一个非空值的行
                    data.append(row_data)
            
            print(f"读取到 {len(data)} 行数据")
            
            # 创建DataFrame
            df = pd.DataFrame(data)
            return df
        else:
            print("Excel文件中没有工作表")
            return None
    except Exception as e:
        print(f"使用原生openpyxl读取文件时出错: {str(e)}")
        return None

def read_excel_file_with_xlrd(file_path):
    """使用xlrd读取Excel文件"""
    print(f"使用xlrd读取文件: {file_path}")
    try:
        # 打开Excel文件
        workbook = xlrd.open_workbook(file_path)
        
        # 获取第一个工作表
        if workbook.nsheets > 0:
            sheet = workbook.sheet_by_index(0)
            print(f"工作表名称: {sheet.name}")
            print(f"行数: {sheet.nrows}, 列数: {sheet.ncols}")
            
            if sheet.nrows <= 1 or sheet.ncols == 0:
                print("没有找到有效数据")
                return None
            
            # 获取列名（假设第一行是表头）
            headers = []
            for col in range(sheet.ncols):
                cell_value = sheet.cell_value(0, col)
                headers.append(cell_value if cell_value else f"Column_{col}")
            
            print(f"列名: {headers}")
            
            # 创建存储数据的列表
            data = []
            
            # 读取数据行
            for row in range(1, sheet.nrows):
                row_data = {}
                all_empty = True
                for col in range(sheet.ncols):
                    cell_value = sheet.cell_value(row, col)
                    if cell_value:
                        all_empty = False
                    row_data[headers[col]] = cell_value
                
                if not all_empty:  # 只添加至少有一个非空值的行
                    data.append(row_data)
            
            print(f"读取到 {len(data)} 行数据")
            
            # 创建DataFrame
            df = pd.DataFrame(data)
            return df
        else:
            print("Excel文件中没有工作表")
            return None
    except Exception as e:
        print(f"使用xlrd读取Excel文件时出错: {str(e)}")
        return None

def read_excel_file(file_path):
    """读取Excel文件，尝试使用不同的引擎和方法"""
    file_size = os.path.getsize(file_path)
    print(f"文件大小: {file_size} 字节")
    
    file_type = read_binary_file_head(file_path)
    
    # 如果是自选人数或保有量文件，优先使用pyexcel尝试读取
    if '自选人数' in file_path or '保有量' in file_path:
        print("检测到自选人数或保有量文件，优先使用pyexcel方法读取")
        df = read_with_pyexcel(file_path)
        if df is not None and len(df) > 0:
            return df
    
    # 1. 然后尝试使用pandas + openpyxl
    try:
        df = pd.read_excel(file_path, engine='openpyxl')
        rows, cols = df.shape
        print(f"成功使用pandas+openpyxl读取文件: {file_path}, 读取到 {rows} 行, {cols} 列")
        
        if rows > 0:
            # 检查是否有实际内容（不只是表头）
            if df.iloc[0].notna().any():
                return df
            else:
                print("数据只包含表头或全为空值，尝试其他方法")
        else:
            print("读取到0行数据，尝试其他方法")
    except Exception as e:
        print(f"使用pandas+openpyxl读取失败: {str(e)}")
    
    # 2. 然后尝试使用原生openpyxl直接读取工作表单元格
    if file_type == "xlsx":
        df = read_with_raw_openpyxl(file_path)
        if df is not None and len(df) > 0:
            return df
    
    # 3. 然后尝试使用pandas + xlrd
    try:
        df = pd.read_excel(file_path, engine='xlrd')
        rows, cols = df.shape
        print(f"成功使用pandas+xlrd引擎读取文件: {file_path}, 读取到 {rows} 行, {cols} 列")
        
        if rows > 0:
            return df
        else:
            print("读取到0行数据，尝试其他方法")
    except Exception as e:
        print(f"使用pandas+xlrd引擎读取失败: {str(e)}")
    
    # 4. 最后尝试手动使用xlrd
    if file_type == "xls":
        df = read_excel_file_with_xlrd(file_path)
        if df is not None and len(df) > 0:
            return df
    
    # 5. 如果确实文件存在但内容为空，返回None
    print(f"无法从文件 {file_path} 中读取有效数据")
    return None

def import_attention_data(db, force_reimport=False):
    """导入ETF自选历史数据"""
    print("\n开始导入ETF自选历史数据...")
    
    # 查找所有自选人数文件
    data_dir = 'data'
    attention_files = [f for f in os.listdir(data_dir) if '自选人数' in f and f.endswith('.xlsx')]
    attention_files.sort()  # 按文件名排序
    
    print(f"找到 {len(attention_files)} 个ETF自选数据文件")
    
    # 获取已有的自选历史数据日期
    cursor = db.conn.cursor()
    cursor.execute("SELECT DISTINCT date FROM etf_attention_history")
    existing_dates = [row[0] for row in cursor.fetchall()]
    print(f"已有自选历史数据日期: {existing_dates}")
    
    # 导入文件
    imported_files = []
    for filename in attention_files:
        file_path = os.path.join(data_dir, filename)
        date = extract_date_from_filename(filename)
        if not date:
            print(f"无法从文件名 {filename} 提取日期，跳过")
            continue
        
        # 检查该日期的数据是否已经存在
        if date in existing_dates and not force_reimport:
            print(f"日期 {date} 的ETF自选历史数据已存在，跳过导入")
            continue
        
        print(f"\n处理ETF自选数据文件: {file_path}, 日期: {date}")
        
        # 读取Excel文件
        df = read_excel_file(file_path)
        if df is None or len(df) == 0:
            print(f"文件 {filename} 读取失败或没有数据")
            continue
        
        print(f"文件 {filename} 的列名: {df.columns.tolist()}")
        
        # 查找代码列和自选人数列
        code_column = None
        attention_column = None
        
        # 检查是否已经有处理过的列
        if 'code' in df.columns and 'attention_count' in df.columns:
            code_column = 'code'
            attention_column = 'attention_count'
            print("找到预处理的code和attention_count列")
        else:
            # 尝试找到代码列
            possible_code_columns = ['标的代码', '代码', 'code', '基金代码', 'ETF代码']
            for col in possible_code_columns:
                if col in df.columns:
                    code_column = col
                    break
            
            # 尝试找到自选人数列
            possible_attention_columns = ['加自选人数', '自选人数', '关注人数', 'attention_count', '人数']
            for col in possible_attention_columns:
                if col in df.columns:
                    attention_column = col
                    break
        
        if code_column is None or attention_column is None:
            print(f"在文件 {filename} 中找不到必要的列（代码列和自选人数列）")
            continue
        
        print(f"使用列: 代码列={code_column}, 自选人数列={attention_column}")
        
        # 准备数据
        try:
            if 'code' not in df.columns:
                # 确保代码列是字符串并去除空白
                df['code'] = df[code_column].astype(str).str.strip()
                
                # 提取6位数字代码
                df['code'] = df['code'].str.extract(r'(\d{6})').fillna('')
            
            # 过滤掉无效代码
            df = df[df['code'] != '']
            
            if 'attention_count' not in df.columns:
                # 转换自选人数为数字
                df['attention_count'] = pd.to_numeric(df[attention_column], errors='coerce').fillna(0).astype(int)
            
            # 添加日期和更新时间
            current_time = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            
            # 过滤有效数据
            valid_data = df[['code', 'attention_count']].dropna(subset=['code'])
            
            if len(valid_data) > 0:
                # 如果是重新导入，先删除已存在的数据
                if date in existing_dates and force_reimport:
                    cursor.execute("DELETE FROM etf_attention_history WHERE date = ?", (date,))
                    print(f"已删除日期 {date} 的现有数据以重新导入")
                
                # 插入数据
                count = 0
                for _, row in valid_data.iterrows():
                    try:
                        cursor.execute("""
                            INSERT INTO etf_attention_history (code, attention_count, date, update_time)
                            VALUES (?, ?, ?, ?)
                        """, (row['code'], row['attention_count'], date, current_time))
                        count += 1
                    except Exception as e:
                        print(f"插入自选数据时出错: {str(e)}")
                
                db.conn.commit()
                print(f"成功导入 {count} 条 {date} 的ETF自选历史数据")
                imported_files.append((filename, date, count))
                
                # 如果是新日期，添加到已存在日期列表中
                if date not in existing_dates:
                    existing_dates.append(date)
            else:
                print(f"文件 {filename} 中没有找到有效数据")
        
        except Exception as e:
            print(f"处理ETF自选数据文件 {file_path} 时出错: {str(e)}")
            import traceback
            traceback.print_exc()
    
    # 查询并显示最终结果
    cursor.execute("SELECT date, COUNT(*) FROM etf_attention_history GROUP BY date ORDER BY date")
    attention_dates = cursor.fetchall()
    print("\nETF自选历史数据日期统计:")
    for date, count in attention_dates:
        print(f"  - {date}: {count}条记录")
    
    return imported_files

def import_holders_data(db, force_reimport=False):
    """导入ETF持有人历史数据"""
    print("\n开始导入ETF持有人历史数据...")
    
    # 查找所有持有人数据文件
    data_dir = 'data'
    holders_files = [f for f in os.listdir(data_dir) if ('保有量' in f or '持有人' in f) and f.endswith('.xlsx')]
    holders_files.sort()  # 按文件名排序
    
    print(f"找到 {len(holders_files)} 个ETF持有人数据文件")
    
    # 获取已有的持有人历史数据日期
    cursor = db.conn.cursor()
    cursor.execute("SELECT DISTINCT date FROM etf_holders_history")
    existing_dates = [row[0] for row in cursor.fetchall()]
    print(f"已有持有人历史数据日期: {existing_dates}")
    
    # 导入文件
    imported_files = []
    for filename in holders_files:
        file_path = os.path.join(data_dir, filename)
        date = extract_date_from_filename(filename)
        if not date:
            print(f"无法从文件名 {filename} 提取日期，跳过")
            continue
        
        # 检查该日期的数据是否已经存在
        if date in existing_dates and not force_reimport:
            print(f"日期 {date} 的ETF持有人历史数据已存在，跳过导入")
            continue
        
        print(f"\n处理ETF持有人数据文件: {file_path}, 日期: {date}")
        
        # 读取Excel文件
        df = read_excel_file(file_path)
        if df is None or len(df) == 0:
            print(f"文件 {filename} 读取失败或没有数据")
            continue
        
        print(f"文件 {filename} 的列名: {df.columns.tolist()}")
        
        # 查找代码列、持有人数列、持仓份额列和持仓价值列
        code_column = None
        holder_count_column = None
        holding_amount_column = None
        holding_value_column = None
        
        # 检查是否已经有处理过的列
        if 'code' in df.columns and 'holder_count' in df.columns and 'holding_amount' in df.columns and 'holding_value' in df.columns:
            code_column = 'code'
            holder_count_column = 'holder_count'
            holding_amount_column = 'holding_amount'
            holding_value_column = 'holding_value'
            print("找到预处理的code、holder_count、holding_amount和holding_value列")
        else:
            # 尝试找到代码列
            possible_code_columns = ['标的代码', '代码', 'code', '基金代码', 'ETF代码']
            for col in possible_code_columns:
                if col in df.columns:
                    code_column = col
                    break
            
            # 尝试找到持有人数列
            possible_holder_count_columns = ['客户数', '持仓客户数', '持有人数', '持有用户数', 'holder_count', '人数']
            for col in possible_holder_count_columns:
                if col in df.columns:
                    holder_count_column = col
                    break
            
            # 尝试找到持仓份额列
            possible_holding_amount_columns = ['持仓份额', '持有份额', 'holding_amount', '份额']
            for col in possible_holding_amount_columns:
                if col in df.columns:
                    holding_amount_column = col
                    break
            
            # 尝试找到持仓价值列
            possible_holding_value_columns = ['持仓市值', '持有市值', 'holding_value', '市值', '保有量', '持有金额', '持仓价值']
            for col in possible_holding_value_columns:
                if col in df.columns:
                    holding_value_column = col
                    break
        
        if code_column is None or holder_count_column is None or (holding_amount_column is None and holding_value_column is None):
            print(f"在文件 {filename} 中找不到必要的列（代码列、持有人数列和至少一个持仓列）")
            continue
        
        print(f"使用列: 代码列={code_column}, 持有人数列={holder_count_column}, 持仓份额列={holding_amount_column}, 持仓价值列={holding_value_column}")
        
        # 准备数据
        try:
            if 'code' not in df.columns:
                # 确保代码列是字符串并去除空白
                df['code'] = df[code_column].astype(str).str.strip()
                
                # 提取6位数字代码
                df['code'] = df['code'].str.extract(r'(\d{6})').fillna('')
            
            # 过滤掉无效代码
            df = df[df['code'] != '']
            
            if 'holder_count' not in df.columns:
                # 转换持有人数为整数
                df['holder_count'] = pd.to_numeric(df[holder_count_column], errors='coerce').fillna(0).astype(int)
            
            if 'holding_amount' not in df.columns and holding_amount_column:
                # 转换持仓份额为浮点数
                df['holding_amount'] = pd.to_numeric(df[holding_amount_column], errors='coerce').fillna(0)
            
            if 'holding_value' not in df.columns and holding_value_column:
                # 转换持仓价值为浮点数
                df['holding_value'] = pd.to_numeric(df[holding_value_column], errors='coerce').fillna(0)
            
            # 添加日期和更新时间
            current_time = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            
            # 过滤有效数据
            valid_data = df[['code', 'holder_count', 'holding_amount', 'holding_value']].dropna(subset=['code'])
            
            if len(valid_data) > 0:
                # 如果是重新导入，先删除已存在的数据
                if date in existing_dates and force_reimport:
                    cursor.execute("DELETE FROM etf_holders_history WHERE date = ?", (date,))
                    print(f"已删除日期 {date} 的现有数据以重新导入")
                
                # 插入数据
                count = 0
                for _, row in valid_data.iterrows():
                    try:
                        cursor.execute("""
                            INSERT INTO etf_holders_history (code, holder_count, holding_amount, holding_value, date, update_time)
                            VALUES (?, ?, ?, ?, ?, ?)
                        """, (row['code'], row['holder_count'], row.get('holding_amount', 0), row.get('holding_value', 0), date, current_time))
                        count += 1
                    except Exception as e:
                        print(f"插入持有人数据时出错: {str(e)}")
                
                db.conn.commit()
                print(f"成功导入 {count} 条 {date} 的ETF持有人历史数据")
                imported_files.append((filename, date, count))
                
                # 如果是新日期，添加到已存在日期列表中
                if date not in existing_dates:
                    existing_dates.append(date)
            else:
                print(f"文件 {filename} 中没有找到有效数据")
        
        except Exception as e:
            print(f"处理ETF持有人数据文件 {file_path} 时出错: {str(e)}")
            import traceback
            traceback.print_exc()
    
    # 查询并显示最终结果
    cursor.execute("SELECT date, COUNT(*) FROM etf_holders_history GROUP BY date ORDER BY date")
    holders_dates = cursor.fetchall()
    print("\nETF持有人历史数据日期统计:")
    for date, count in holders_dates:
        print(f"  - {date}: {count}条记录")
    
    return imported_files

def main():
    """主函数"""
    # 连接数据库
    db = Database()
    
    # 确保历史表存在
    cursor = db.conn.cursor()
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS etf_attention_history (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            code TEXT,
            attention_count INTEGER,
            date TEXT,
            update_time TIMESTAMP
        )
    """)
    
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS etf_holders_history (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            code TEXT,
            holder_count INTEGER,
            holding_amount REAL,
            holding_value REAL,
            date TEXT,
            update_time TIMESTAMP,
            UNIQUE(code, date)
        )
    """)
    
    # 导入ETF自选历史数据
    force_reimport = True  # 设置为True将重新导入已存在的数据
    imported_attention_files = import_attention_data(db, force_reimport)
    
    # 导入ETF持有人历史数据
    imported_holders_files = import_holders_data(db, force_reimport)
    
    print("\n导入ETF历史数据完成")
    print(f"导入ETF自选历史数据文件数: {len(imported_attention_files)}")
    print(f"导入ETF持有人历史数据文件数: {len(imported_holders_files)}")

if __name__ == "__main__":
    main() 