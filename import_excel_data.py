#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
ETF数据导入工具

这个脚本用于将各种Excel数据导入到ETF数据库中。
它提供了一个统一的入口点，可以选择导入不同类型的数据。
"""

import os
import glob
import json
import pandas as pd
import sqlite3
from datetime import datetime
import logging
import sys
import argparse
from database.models import Database
import re
from utils.etf_code import normalize_etf_code

# 配置日志
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.StreamHandler(sys.stdout),
        logging.FileHandler('etf_import.log')
    ]
)
logger = logging.getLogger(__name__)

# 数据目录
DATA_DIR = 'data'
os.makedirs(DATA_DIR, exist_ok=True)

def find_latest_file(pattern):
    """查找匹配模式的最新文件"""
    files = glob.glob(os.path.join(DATA_DIR, pattern))
    if not files:
        return None
    
    # 从文件名中提取日期，并选择日期最新的文件
    latest_file = None
    latest_date = None
    
    for file in files:
        filename = os.path.basename(file)
        date_match = re.search(r'(\d{8})', filename)
        if date_match:
            file_date = date_match.group(1)
            if latest_date is None or file_date > latest_date:
                latest_date = file_date
                latest_file = file
    
    return latest_file

def save_column_mapping():
    """保存列名映射到JSON文件"""
    column_mapping = {
        'etf_info': {
            '证券代码': 'code',
            '证券简称': 'name',
            '业绩比较基准': 'benchmark',
            '成立年限\n[单位] 年': 'years_since_establishment',
            '基金上市地点': 'listing_location',
            '基金成立日': 'establishment_date',
            '基金管理人': 'fund_manager',
            '管理费率\n[单位] %': 'management_fee_rate',
            '托管费率\n[单位] %': 'custodian_fee_rate',
            '指数使用费率': 'index_usage_fee_rate',
            '跟踪指数代码': 'tracking_index_code',
            '日均跟踪偏离度阈值(业绩基准)\n[单位] %': 'tracking_deviation_threshold',
            '年化跟踪误差阈值(业绩基准)\n[单位] %': 'tracking_error_threshold',
            '跟踪误差\n[起始交易日期] S_cal_date(enddate,-52,W,0)\n[ 截止交易日期] 最新收盘日\n[计算周期] 日\n[收益率计算方法] 普通收益率\n[标的指数] 上证综合指数\n[单位] %': 'tracking_error',
            '跟踪误差(跟踪指数)\n[起始交易日期] S_cal_date(enddate,-52,W,0)\n[截止交易日期] 最新收盘日\n[计算周期] 日\n[ 收益率计算方法] 普通收益率\n[单位] %': 'tracking_error_index',
            '信息比率(年化)\n[起始交易日期] S_cal_date(enddate,-52,W,0)\n[截止交易日期] 最新收盘日\n[计算周期] 日\n[收益率计算方法] 普通收益率\n[无风险收益率] 一年定存利率（税前）\n[标的指数] 上证综合指数': 'information_ratio',
            '跟踪误差(年化)\n[起始交易日期] S_cal_date(enddate,-52,W,0)\n[截止交易日期] 最新收盘日\n[计算周期] 日\n[收益率计算方法] 普通收益率\n[标的指数] 上证综合指数\n[单位] %': 'tracking_error_annualized',
            '基金规模(合计)\n[交易日期] S_cal_date(now(),0,D,0)\n[单位] 亿元': 'fund_size',
            '基金份额持有人户数\n[报告期] 20240630\n[单位] 户': 'holder_count',
            '基金份额持有人户数(合计)\n[报告期] 20240630\n[单位] 户': 'total_holder_count',
            '区间日均成交额\n[起始交易日期] S_cal_date(enddate,-1,M,0)\n[截止交易日期] 最新收盘日\n[单位] 亿元': 'avg_daily_volume',
            '基金管理人简称': 'fund_manager_abbr',
            '基金场内简称': 'fund_exchange_abbr',
            '跟踪指数名称': 'tracking_index_name',
            '月成交额\n[交易日期] 最新收盘日\n[单位] 百万元': 'monthly_volume',
            '涨跌幅\n[交易日期] 最新收盘日\n[单位] %': 'price_change_rate',
            '换手率\n[交易日期] 最新收盘日\n[单位] %': 'turnover_rate',
            '成交额\n[交易日期] 最新收盘日\n[单位] 亿元': 'daily_volume',
            '成交笔数\n[交易日期] 最新收盘日\n[单位] 笔': 'transaction_count',
            '总市值\n[交易日期] 最新收盘日\n[单位] 亿元': 'market_value'
        },
        'etf_price': {
            '证券代码': 'code',
            '涨跌幅\n[交易日期] 最新收盘日\n[单位] %': 'price_change_rate',
            '换手率\n[交易日期] 最新收盘日\n[单位] %': 'turnover_rate',
            '成交额\n[交易日期] 最新收盘日\n[单位] 亿元': 'daily_volume',
            '成交笔数\n[交易日期] 最新收盘日\n[单位] 笔': 'transaction_count',
            '总市值\n[交易日期] 最新收盘日\n[单位] 亿元': 'market_value'
        },
        'etf_attention': {
            '标的代码': 'code',
            '加自选人数': 'attention_count'
        },
        'etf_holders': {
            '证券代码': 'code',
            '持有人数': 'holder_count',
            '持有市值': 'holding_value',
            '持仓价值': 'holding_value',
            '客户保有量（元）': 'holding_value'
        }
    }
    
    # 保存列名映射到JSON文件
    mapping_file = os.path.join(DATA_DIR, 'etf_column_names.json')
    with open(mapping_file, 'w', encoding='utf-8') as f:
        json.dump(column_mapping, f, ensure_ascii=False, indent=4)
    
    logger.info(f"列名映射已保存到: {mapping_file}")
    return column_mapping

def import_etf_info():
    """导入ETF基本信息"""
    try:
        logger.info("开始导入ETF基本信息...")
        
        # 查找最新的ETF数据文件
        etf_file = find_latest_file("ETF_DATA_*.xlsx")
        if not etf_file:
            logger.error("未找到ETF数据文件")
            return False
        
        logger.info(f"使用文件: {etf_file}")
        
        # 从文件名中提取日期
        filename = os.path.basename(etf_file)
        date_match = re.search(r'(\d{8})', filename)
        if date_match:
            data_date = date_match.group(1)
            formatted_date = f"{data_date[:4]}-{data_date[4:6]}-{data_date[6:8]}"
            logger.info(f"从文件名中提取的日期: {formatted_date}")
        else:
            # 如果无法从文件名中提取日期，使用当前日期，但记录警告
            formatted_date = datetime.now().strftime('%Y-%m-%d')
            logger.warning(f"无法从文件名中提取日期，使用当前日期: {formatted_date}")
        
        # 读取Excel文件
        df = pd.read_excel(etf_file, engine='openpyxl')
        
        # 添加日期字段 - 使用从文件名提取的日期，而非当前日期
        df['date'] = formatted_date
        
        # 创建数据库连接
        db = Database()
        
        # 保存ETF基本信息
        if db.save_etf_info(df):
            logger.info(f"成功导入ETF基本信息，共{len(df)}条记录")
            return True
        else:
            logger.error("导入ETF基本信息失败")
            return False
        
    except Exception as e:
        logger.error(f"导入ETF基本信息时出错: {str(e)}", exc_info=True)
        return False

def import_etf_price():
    """导入ETF价格数据"""
    try:
        logger.info("开始导入ETF价格数据...")

        # 查找最新的ETF数据文件
        etf_file = find_latest_file("ETF_DATA_*.xlsx")
        if not etf_file:
            logger.error("未找到ETF数据文件")
            return False

        logger.info(f"使用文件: {etf_file}")

        # 从文件名中提取日期
        filename = os.path.basename(etf_file)
        date_match = re.search(r'(\d{8})', filename)
        if date_match:
            data_date = date_match.group(1)
            formatted_date = f"{data_date[:4]}-{data_date[4:6]}-{data_date[6:8]}"
            logger.info(f"从文件名中提取的日期: {formatted_date}")
        else:
            # 如果无法从文件名中提取日期，使用当前日期，但记录警告
            formatted_date = datetime.now().strftime('%Y-%m-%d')
            logger.warning(f"无法从文件名中提取日期，使用当前日期: {formatted_date}")

        # 读取Excel文件
        df = pd.read_excel(etf_file, engine='openpyxl')
        
        # 标准化列名
        df.columns = [str(col).strip() for col in df.columns]

        logger.info(f"原始列名: {list(df.columns)}")

        # 定义需要的列名和映射关系
        new_column_map = {
            '证券代码': 'code',
            '涨跌幅\n[交易日期] 最新收盘日\n[单位] %': 'change_rate',
            '换手率\n[交易日期] 最新收盘日\n[单位] %': 'turnover_rate',
            '成交额\n[交易日期] 最新收盘日\n[单位] 亿元': 'amount',
            '成交笔数\n[交易日期] 最新收盘日\n[单位] 笔': 'transaction_count',
            '总市值\n[交易日期] 最新收盘日\n[单位] 亿元': 'total_market_value',
            '收盘价\n[交易日期] 最新收盘日\n[复权方式] 不复权\n[单位] 元': 'close_price',
            '开盘价\n[交易日期] 最新收盘日\n[复权方式] 不复权\n[单位] 元': 'open_price',
            '最高价\n[交易日期] 最新收盘日\n[复权方式] 不复权\n[单位] 元': 'high_price',
            '最低价\n[交易日期] 最新收盘日\n[复权方式] 不复权\n[单位] 元': 'low_price',
            '振幅\n[交易日期] 最新收盘日\n[单位] %': 'amplitude',
            '升贴水\n[交易日期] 最新收盘日\n[单位] 元': 'premium_discount',
            '升贴水率\n[交易日期] 最新收盘日\n[单位] %': 'premium_discount_rate'
        }

        # 识别并匹配列名（考虑换行符和空格的变化）
        matched_columns = {}
        for excel_col in df.columns:
            excel_col_clean = str(excel_col).strip().replace('\n', '')
            for pattern_col, db_col in new_column_map.items():
                pattern_col_clean = pattern_col.replace('\n', '')
                if excel_col_clean == pattern_col_clean or excel_col_clean.startswith(pattern_col_clean):
                    matched_columns[excel_col] = db_col
                    break

        # 检查是否找到了证券代码列
        if '证券代码' not in matched_columns.keys() and not any(col.startswith('证券代码') for col in df.columns):
            logger.error("未找到证券代码列，无法继续导入")
            return False

        # 重命名识别到的列
        logger.info(f"匹配到的列映射: {matched_columns}")
        df = df.rename(columns=matched_columns)
        
        # 确保code列存在
        if 'code' not in df.columns:
            # 尝试查找以"证券代码"开头的列
            code_cols = [col for col in df.columns if str(col).startswith('证券代码')]
            if code_cols:
                df = df.rename(columns={code_cols[0]: 'code'})
            else:
                logger.error("未找到证券代码列，无法继续导入")
                return False

        # 标准化ETF代码
        df['code'] = df['code'].apply(normalize_etf_code)
        
        # 添加日期列
        df['date'] = formatted_date
        
        # 添加更新时间列
        df['update_time'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

        # 检查是否有足够的数据列
        min_required_columns = ['code', 'date', 'update_time']
        price_data_columns = [col for col in df.columns if col not in ['code', 'date', 'update_time']]
        
        if len(price_data_columns) < 2:
            logger.error(f"ETF价格数据文件缺少足够的数据列，只找到: {price_data_columns}")
            return False
            
        logger.info(f"最终列名: {list(df.columns)}")
        logger.info(f"数据示例:\n{df.head()}")

        # 创建数据库连接
        db = Database()

        # 保存到数据库
        if db.save_etf_price(df):
            logger.info(f"成功导入ETF价格数据，共{len(df)}条记录")
            return True
        else:
            logger.error("导入ETF价格数据失败")
            return False
            
    except Exception as e:
        logger.error(f"导入ETF价格数据时出错: {str(e)}", exc_info=True)
        return False

def import_etf_holders():
    """导入ETF持有人数据"""
    try:
        logger.info("开始导入ETF持有人数据...")
        
        # 查找最新的ETF持有人数据文件
        holder_file = find_latest_file("客户ETF保有量*.xlsx")
        if not holder_file:
            logger.error("未找到ETF持有人数据文件")
            return False
        
        logger.info(f"使用文件: {holder_file}")
        
        # 从文件名中提取日期
        filename = os.path.basename(holder_file)
        date_match = re.search(r'(\d{8})', filename)
        if date_match:
            data_date = date_match.group(1)
            formatted_date = f"{data_date[:4]}-{data_date[4:6]}-{data_date[6:8]}"
            logger.info(f"从文件名中提取的日期: {formatted_date}")
        else:
            # 如果无法从文件名中提取日期，使用当前日期，但记录警告
            formatted_date = datetime.now().strftime('%Y-%m-%d')
            logger.warning(f"无法从文件名中提取日期，使用当前日期: {formatted_date}")
        
        # 尝试列出所有工作表
        xl = pd.ExcelFile(holder_file)
        sheet_names = xl.sheet_names
        logger.info(f"文件中的工作表: {sheet_names}")
        
        # 如果只有一个工作表，直接读取
        if len(sheet_names) == 1:
            # 首先尝试直接读取
            df = pd.read_excel(holder_file, engine='openpyxl')
        elif '数据表' in sheet_names:
            # 尝试读取名为"数据表"的工作表
            df = pd.read_excel(holder_file, sheet_name='数据表', engine='openpyxl')
        elif 'Sheet1' in sheet_names:
            # 尝试读取Sheet1
            df = pd.read_excel(holder_file, sheet_name='Sheet1', engine='openpyxl')
        else:
            # 尝试读取第一个工作表
            df = pd.read_excel(holder_file, sheet_name=sheet_names[0], engine='openpyxl')
        
        # 打印列名和数据样本，帮助诊断
        logger.info(f"原始列名: {list(df.columns)}")
        logger.info(f"数据形状: {df.shape}")
        logger.info(f"数据前5行:\n{df.head()}")
        
        # 如果数据帧为空，尝试其他读取方式
        if df.empty or len(df.columns) <= 1:
            logger.warning("标准读取方式返回空数据或只有单列，尝试读取所有单元格...")
            # 使用更底层的方式读取
            import openpyxl
            wb = openpyxl.load_workbook(holder_file)
            
            # 尝试所有工作表
            for sheet_name in wb.sheetnames:
                logger.info(f"尝试读取工作表: {sheet_name}")
                ws = wb[sheet_name]
                
                # 获取表格范围
                data = []
                # 从第一行开始，假设第一行是标题
                for row in ws.iter_rows(min_row=1, values_only=True):
                    if any(row):  # 只添加非空行
                        data.append(row)
                
                if data:
                    # 检查第一行是否包含我们需要的列标题
                    first_row = data[0]
                    logger.info(f"找到的标题行: {first_row}")
                    
                    # 检查是否包含必要的列
                    if '标的代码' in first_row and '持仓客户数' in first_row and '持仓市值' in first_row:
                        df = pd.DataFrame(data[1:], columns=data[0])
                        logger.info(f"成功读取数据，形状: {df.shape}")
                        break
                    
            if df.empty or len(df.columns) <= 1:
                logger.error("所有读取方式都失败，无法获取必要的数据")
                return False
        
        # 确保所需列存在
        required_columns = ['标的代码', '持仓客户数', '持仓市值']
        if '持仓份额' in df.columns:
            required_columns.append('持仓份额')
            
        missing_columns = [col for col in required_columns if col not in df.columns]
        
        if missing_columns:
            logger.error(f"Excel文件缺少以下列: {missing_columns}")
            logger.error(f"实际列名: {list(df.columns)}")
            return False
        
        # 重命名列
        rename_map = {
            '标的代码': 'code',
            '持仓客户数': 'holder_count',
            '持仓市值': 'holding_value'
        }
        
        if '持仓份额' in df.columns:
            rename_map['持仓份额'] = 'holding_amount'
            
        df = df.rename(columns=rename_map)
        
        # 标准化ETF代码
        df['code'] = df['code'].apply(normalize_etf_code)
        
        # 转换数据类型
        df['holder_count'] = pd.to_numeric(df['holder_count'], errors='coerce').fillna(0).astype(int)
        df['holding_value'] = pd.to_numeric(df['holding_value'], errors='coerce').fillna(0).astype(float)
        
        if 'holding_amount' in df.columns:
            df['holding_amount'] = pd.to_numeric(df['holding_amount'], errors='coerce').fillna(0).astype(float)
        else:
            # 如果没有持仓份额列，将字段设为0而不是使用持仓市值
            df['holding_amount'] = 0
        
        # 添加日期字段
        df['date'] = formatted_date
        
        # 打印处理后的数据帮助诊断
        logger.info(f"处理后列名: {list(df.columns)}")
        logger.info(f"处理后数据前5行:\n{df.head()}")
        
        # 创建数据库连接
        db = Database()
        
        # 保存到数据库
        if db.save_etf_holders(df):
            logger.info(f"成功导入ETF持有人数据，共{len(df)}条记录")
            return True
        else:
            logger.error("导入ETF持有人数据失败")
            return False
        
    except Exception as e:
        logger.error(f"导入ETF持有人数据时出错: {str(e)}", exc_info=True)
        return False
    
def import_etf_attention():
    """导入ETF自选数据"""
    try:
        logger.info("开始导入ETF自选数据...")
        
        # 查找最新的ETF自选数据文件
        attention_file = find_latest_file("客户ETF自选人数*.xlsx")
        if not attention_file:
            logger.error("未找到ETF自选数据文件")
            return False

        logger.info(f"使用文件: {attention_file}")
        
        # 从文件名中提取日期
        filename = os.path.basename(attention_file)
        date_match = re.search(r'(\d{8})', filename)
        if date_match:
            data_date = date_match.group(1)
            formatted_date = f"{data_date[:4]}-{data_date[4:6]}-{data_date[6:8]}"
            logger.info(f"从文件名中提取的日期: {formatted_date}")
        else:
            # 如果无法从文件名中提取日期，使用当前日期，但记录警告
            formatted_date = datetime.now().strftime('%Y-%m-%d')
            logger.warning(f"无法从文件名中提取日期，使用当前日期: {formatted_date}")
        
        # 尝试列出所有工作表
        xl = pd.ExcelFile(attention_file)
        sheet_names = xl.sheet_names
        logger.info(f"文件中的工作表: {sheet_names}")
        
        # 如果只有一个工作表，直接读取
        if len(sheet_names) == 1:
            # 首先尝试直接读取
            df = pd.read_excel(attention_file, engine='openpyxl')
        elif '客户ETF自选人数' in sheet_names:
            # 尝试读取名为"客户ETF自选人数"的工作表
            df = pd.read_excel(attention_file, sheet_name='客户ETF自选人数', engine='openpyxl')
        elif 'Sheet1' in sheet_names:
            # 尝试读取Sheet1
            df = pd.read_excel(attention_file, sheet_name='Sheet1', engine='openpyxl')
        else:
            # 尝试读取第一个工作表
            df = pd.read_excel(attention_file, sheet_name=sheet_names[0], engine='openpyxl')
        
        # 打印列名和数据样本，帮助诊断
        logger.info(f"原始列名: {list(df.columns)}")
        logger.info(f"数据形状: {df.shape}")
        logger.info(f"数据前5行:\n{df.head()}")
        
        # 如果数据帧为空，尝试其他读取方式
        if df.empty or len(df.columns) <= 1:
            logger.warning("标准读取方式返回空数据或只有单列，尝试读取所有单元格...")
            # 使用更底层的方式读取
            import openpyxl
            wb = openpyxl.load_workbook(attention_file)
            
            # 尝试所有工作表
            for sheet_name in wb.sheetnames:
                logger.info(f"尝试读取工作表: {sheet_name}")
                ws = wb[sheet_name]
                
                # 获取表格范围
                data = []
                # 从第一行开始，假设第一行是标题
                for row in ws.iter_rows(min_row=1, values_only=True):
                    if any(row):  # 只添加非空行
                        data.append(row)
                
                if data:
                    # 检查第一行是否包含我们需要的列标题
                    first_row = data[0]
                    logger.info(f"找到的标题行: {first_row}")
                    
                    # 检查是否包含必要的列
                    if '标的代码' in first_row and '加自选人数' in first_row:
                        df = pd.DataFrame(data[1:], columns=data[0])
                        logger.info(f"成功读取数据，形状: {df.shape}")
                        break
                    
            if df.empty or len(df.columns) <= 1:
                logger.error("所有读取方式都失败，无法获取必要的数据")
                return False
        
        # 标准化列名
        df.columns = [str(col).strip() for col in df.columns]
        
        # 检查所需列是否存在
        required_columns = ['标的代码', '加自选人数']
        
        # 检查实际列名中是否有与所需列名模糊匹配的
        actual_columns = list(df.columns)
        logger.info(f"标准化后列名: {actual_columns}")
        
        # 查找可能的代码列
        code_columns = [col for col in actual_columns if '代码' in col]
        logger.info(f"找到的代码列: {code_columns}")
        
        # 查找可能的自选列
        attention_columns = [col for col in actual_columns if '自选' in col or '加' in col or '人数' in col]
        logger.info(f"找到的自选列: {attention_columns}")
        
        # 如果找不到精确匹配的列名，尝试使用模糊匹配的列名
        if '标的代码' not in actual_columns and code_columns:
            logger.info(f"使用替代代码列: {code_columns[0]}")
            df = df.rename(columns={code_columns[0]: '标的代码'})
            actual_columns = list(df.columns)
        
        if '加自选人数' not in actual_columns and attention_columns:
            logger.info(f"使用替代自选人数列: {attention_columns[0]}")
            df = df.rename(columns={attention_columns[0]: '加自选人数'})
            actual_columns = list(df.columns)
        
        # 再次检查所需列是否存在
        missing_columns = [col for col in required_columns if col not in actual_columns]
        
        if missing_columns:
            logger.error(f"未找到代码列")
            logger.error(f"文件中的所有列: {actual_columns}")
            return False
        
        # 重命名列
        rename_map = {
            '标的代码': 'code',
            '加自选人数': 'attention_count'
        }
        df = df.rename(columns=rename_map)
        
        # 标准化ETF代码
        df['code'] = df['code'].apply(normalize_etf_code)
        
        # 转换数据类型
        df['attention_count'] = pd.to_numeric(df['attention_count'], errors='coerce').fillna(0).astype(int)
        
        # 添加日期字段
        df['date'] = formatted_date
        
        # 打印处理后的数据帮助诊断
        logger.info(f"处理后列名: {list(df.columns)}")
        logger.info(f"处理后数据前5行:\n{df.head()}")
        
        # 创建数据库连接
        db = Database()
        
        # 保存到数据库
        if db.save_etf_attention(df):
            logger.info(f"成功导入ETF自选数据，共{len(df)}条记录")
            return True
        else:
            logger.error("导入ETF自选数据失败")
            return False
    
    except Exception as e:
        logger.error(f"导入ETF自选数据时出错: {str(e)}", exc_info=True)
        return False

def import_etf_index_classification():
    """导入ETF指数分类数据"""
    try:
        logger.info("开始导入ETF指数分类数据...")
        
        # 查找最新的ETF指数分类数据文件
        classification_file = find_latest_file("ETF-Index-Classification_*.xlsx")
        if not classification_file:
            logger.error("未找到ETF指数分类数据文件")
            return False
        
        logger.info(f"使用文件: {classification_file}")
        
        # 读取Excel文件
        df = pd.read_excel(classification_file, engine='openpyxl')
        
        # 创建数据库连接
        db = Database()
        
        # 保存到数据库
        if db.save_etf_index_classification(df):
            logger.info(f"成功导入ETF指数分类数据，共{len(df)}条记录")
            return True
        else:
            logger.error("导入ETF指数分类数据失败")
            return False
    
    except Exception as e:
        logger.error(f"导入ETF指数分类数据时出错: {str(e)}", exc_info=True)
        return False

def import_etf_business():
    """导入ETF商务协议数据"""
    try:
        logger.info("开始导入ETF商务协议数据...")
        
        # 查找最新的ETF商务协议数据文件
        business_file = find_latest_file("ETF单产品商务协议*.xlsx")
        if not business_file:
            logger.error("未找到ETF商务协议数据文件")
            return False
        
        logger.info(f"使用文件: {business_file}")
        
        # 读取Excel文件
        df = pd.read_excel(business_file, engine='openpyxl')
        
        # 创建数据库连接
        db = Database()
        
        # 保存到数据库
        if db.save_business_etf(df):
            logger.info(f"成功导入ETF商务协议数据，共{len(df)}条记录")
            return True
        else:
            logger.error("导入ETF商务协议数据失败")
            return False
    
    except Exception as e:
        logger.error(f"导入ETF商务协议数据时出错: {str(e)}", exc_info=True)
        return False

def import_all_data():
    """导入所有ETF数据"""
    logger.info("=== 开始导入所有ETF数据 ===")
    
    # 保存列名映射
    save_column_mapping()
    
    # 导入ETF基本信息
    success_info = import_etf_info()
    
    # 导入ETF价格数据
    success_price = import_etf_price()
    
    # 导入ETF持有人数据
    success_holders = import_etf_holders()
    
    # 导入ETF自选数据
    success_attention = import_etf_attention()
    
    # 导入ETF指数分类数据
    success_classification = import_etf_index_classification()
    
    # 导入ETF商务协议数据
    success_business = import_etf_business()
    
    # 输出导入结果摘要
    logger.info("=== ETF数据导入完成 ===")
    logger.info(f"ETF基本信息: {'成功' if success_info else '失败'}")
    logger.info(f"ETF价格数据: {'成功' if success_price else '失败'}")
    logger.info(f"ETF持有人数据: {'成功' if success_holders else '失败'}")
    logger.info(f"ETF自选数据: {'成功' if success_attention else '失败'}")
    logger.info(f"ETF指数分类数据: {'成功' if success_classification else '失败'}")
    logger.info(f"ETF商务协议数据: {'成功' if success_business else '失败'}")
    
    # 创建数据库连接以获取记录数和更新公司分析数据
    try:
        db = Database()
        conn = db.connect()
        
        etf_info_count = conn.execute("SELECT COUNT(*) FROM etf_info").fetchone()[0]
        etf_price_count = conn.execute("SELECT COUNT(*) FROM etf_price").fetchone()[0]
        etf_holders_count = conn.execute("SELECT COUNT(*) FROM etf_holders").fetchone()[0]
        etf_attention_count = conn.execute("SELECT COUNT(*) FROM etf_attention").fetchone()[0]
        etf_business_count = conn.execute("SELECT COUNT(*) FROM etf_business").fetchone()[0]
        
        logger.info("=== 数据库记录统计 ===")
        logger.info(f"ETF基本信息表: {etf_info_count}条记录")
        logger.info(f"ETF价格数据表: {etf_price_count}条记录")
        logger.info(f"ETF持有人数据表: {etf_holders_count}条记录")
        logger.info(f"ETF自选数据表: {etf_attention_count}条记录")
        logger.info(f"ETF商务协议数据表: {etf_business_count}条记录")
        
        # 更新基金公司分析数据
        logger.info("开始更新基金公司分析数据...")
        if db.update_company_analytics_data():
            logger.info("基金公司分析数据更新成功。")
            company_analytics_count = conn.execute("SELECT COUNT(*) FROM etf_company_analytics").fetchone()[0]
            logger.info(f"基金公司分析表: {company_analytics_count}条记录")
        else:
            logger.error("基金公司分析数据更新失败。")
        
        db.close()
        
    except Exception as e:
        logger.error(f"获取数据库统计信息时出错: {str(e)}")
    
    return all([success_info, success_price, success_holders, success_attention, success_classification, success_business])

def show_menu():
    """显示导入菜单"""
    print("\n===== ETF数据导入工具 =====")
    print("1. 导入所有ETF数据")
    print("2. 仅导入ETF基本信息")
    print("3. 仅导入ETF价格数据")
    print("4. 仅导入ETF持有人数据")
    print("5. 仅导入ETF自选数据")
    print("6. 仅导入ETF指数分类数据")
    print("7. 仅导入ETF商务协议数据")
    print("0. 退出")
    print("=========================")

def main():
    """主函数"""
    # 解析命令行参数
    parser = argparse.ArgumentParser(description='ETF数据导入工具')
    parser.add_argument('--all', action='store_true', help='导入所有数据')
    parser.add_argument('--info', action='store_true', help='导入ETF基本信息')
    parser.add_argument('--price', action='store_true', help='导入ETF价格数据')
    parser.add_argument('--holders', action='store_true', help='导入ETF持有人数据')
    parser.add_argument('--attention', action='store_true', help='导入ETF自选数据')
    parser.add_argument('--classification', action='store_true', help='导入ETF指数分类数据')
    parser.add_argument('--business', action='store_true', help='导入ETF商务协议数据')
    parser.add_argument('--menu', action='store_true', help='显示交互式菜单')
    
    args = parser.parse_args()
    
    # 检查是否有命令行参数
    if args.all or args.info or args.price or args.holders or args.attention or args.classification or args.business:
        # 根据命令行参数导入指定数据
        if args.all:
            import_all_data()
        else:
            if args.info:
                import_etf_info()
            if args.price:
                import_etf_price()
            if args.holders:
                import_etf_holders()
            if args.attention:
                import_etf_attention()
            if args.classification:
                import_etf_index_classification()
            if args.business:
                import_etf_business()
    elif args.menu or len(sys.argv) == 1:
        # 显示交互式菜单
        while True:
            show_menu()
            choice = input("请选择操作 (0-7): ")
            
            if choice == '0':
                print("退出程序")
                break
            elif choice == '1':
                import_all_data()
            elif choice == '2':
                import_etf_info()
            elif choice == '3':
                import_etf_price()
            elif choice == '4':
                import_etf_holders()
            elif choice == '5':
                import_etf_attention()
            elif choice == '6':
                import_etf_index_classification()
            elif choice == '7':
                import_etf_business()
            else:
                print("无效的选择，请重试")
            
            input("\n按Enter键继续...")

if __name__ == "__main__":
    main()