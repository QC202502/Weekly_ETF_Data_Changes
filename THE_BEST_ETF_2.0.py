import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.utils.dataframe import dataframe_to_rows

# 新增版本信息
__version__ = "2.7.4"   
RELEASE_DATE = "2025-03-13"  # 请根据实际发布日期修改

# 文件路径配置
etf_file_path = '/Users/admin/Downloads/ETF_DATA_20250314.xlsx'
classification_file_path = '/Users/admin/Downloads/ETF-Index-Classification_20250314.xlsx'
business_etf_path = '/Users/admin/Downloads/ETF单产品商务协议20250314.xlsx'

# 分类排序规则
CATEGORY_ORDER = {
    '一级分类': ['宽基规模', '主题行业', '策略风格', '商品债券', '跨境', '空白'],
    '二级分类': None,
    '三级分类': None
}

# 读取商务品代码
business_etf_df = pd.read_excel(business_etf_path, engine='openpyxl')
business_etf_codes = business_etf_df['证券代码'].astype(str).str.strip().tolist()

# 预处理ETF数据
etf_data = pd.read_excel(etf_file_path, sheet_name='万得', engine='openpyxl')
etf_data['证券代码_处理'] = etf_data['证券代码'].astype(str).str.split('.').str[0].str.strip()
etf_data = etf_data.rename(columns={
    '管理费率\n[单位] %': '管理费率',
    '托管费率\n[单位] %': '托管费率',
    '月成交额\n[交易日期] 最新收盘日\n[单位] 百万元': '月成交额',
    '基金规模(合计)[交易日期]S_cal_date(now(),0,D,0)[单位]亿元': '规模'
})
etf_data['综合费率'] = etf_data['管理费率'] + etf_data['托管费率']

# 合并分类数据
classification_data = pd.read_excel(classification_file_path, sheet_name='Sheet1', engine='openpyxl')
merged_data = pd.merge(
    etf_data,
    classification_data,
    on='跟踪指数代码',
    how='left',
    suffixes=('', '_分类表')
)


# 在 add_index_statistics 函数中更新商务品合计规模的计算
# 分类统计函数

# 定义一个函数，用于简化基金公司名称
def simplify_fund_company_name(name):
    if pd.isna(name):  # 如果名称为空，直接返回
        return name
    # 找到“基金”的位置，并截取前面的部分
    index = name.find("基金")
    if index != -1:
        return name[:index]  # 返回“基金”前面的部分
    return name  # 如果没有“基金”，返回原名称

# 分类统计函数
def add_index_statistics(group):
    # 筛选商务品
    business_group = group[group['证券代码_处理'].isin(business_etf_codes)]
    
    # 统计商务品数量和合计规模
    business_count = business_group.shape[0]
    business_total_scale = round(business_group['规模'].sum(), 1) if business_count > 0 else 0
    
    # 统计整个跟踪指数的合计规模
    total_scale = round(group['规模'].sum(), 1)
    
    return pd.Series({
        '跟踪ETF数量': group.shape[0],
        '商务品数量': business_count,
        '商务品合计规模': business_total_scale,
        '合计规模': total_scale  # 新增：整个跟踪指数的合计规模
    })

# 分类合计计算函数
def calculate_category_totals(df):
    df['二级分类'] = df['二级分类'].fillna('未分类')
    df['三级分类'] = df['三级分类'].fillna('未分类')

    # 计算二级分类合计（基于整个跟踪指数的合计规模）
    df['二级分类合计'] = df.groupby('二级分类', dropna=False)['合计规模'].transform(
        lambda x: round(x.sum(), 1)
    )

    # 计算三级分类合计（基于整个跟踪指数的合计规模）
    df['三级分类合计'] = df.groupby('三级分类', dropna=False)['合计规模'].transform(
        lambda x: round(x.sum(), 1)
    )

    return df


# 自定义排序函数
def custom_sort(df):
    df['sort_key_1'] = df['一级分类'].map(
        {v: i for i, v in enumerate(CATEGORY_ORDER['一级分类'])}
    ).fillna(len(CATEGORY_ORDER['一级分类']))

    df['二级分类合计'] = df['二级分类合计'].fillna(0)
    df['三级分类合计'] = df['三级分类合计'].fillna(0)

    return df.sort_values(
        by=['sort_key_1', '二级分类合计', '三级分类合计'],
        ascending=[True, False, False]
    ).drop(columns='sort_key_1')


# 在主处理流程中更新 results 的生成逻辑
results = []
for index_code, group in merged_data.groupby('跟踪指数代码'):
    if group.empty:
        continue

    try:
        stats = add_index_statistics(group)

        # 获取最低费率基金
        lowest_fee_group = group[group['综合费率'] == group['综合费率'].min()]
        lowest_fee_fund = lowest_fee_group.loc[lowest_fee_group['月成交额'].idxmax()]

        # 获取最高交易量基金
        highest_volume_fund = group.loc[group['月成交额'].idxmax()]

        # 获取最大规模基金
        largest_scale_fund = group.loc[group['规模'].idxmax()]
        # 判断是否为商务品
        is_lowest_fee_business = lowest_fee_fund['证券代码_处理'] in business_etf_codes
        is_highest_volume_business = highest_volume_fund['证券代码_处理'] in business_etf_codes
        is_largest_scale_business = largest_scale_fund['证券代码_处理'] in business_etf_codes

        # 跟踪指数分级逻辑
        if is_highest_volume_business and is_lowest_fee_business and is_largest_scale_business:
            index_level = '一级'
        elif is_highest_volume_business and is_lowest_fee_business and not is_largest_scale_business:
            index_level = '二级'
        elif is_highest_volume_business and is_largest_scale_business and not is_lowest_fee_business:
            index_level = '三级'
        elif is_lowest_fee_business and is_largest_scale_business and not is_highest_volume_business:
            index_level = '四级'
        elif is_highest_volume_business:
            index_level = '五级'
        elif is_largest_scale_business:
            index_level = '六级'
        elif is_lowest_fee_business:
            index_level = '七级'
        elif stats['商务品数量'] > 0:
            index_level = '八级'
        else:
            index_level = '九级'

        results.append({
            '跟踪指数代码': index_code,
            '跟踪指数名称': group['跟踪指数名称'].iloc[0] if '跟踪指数名称' in group.columns else None,
            '跟踪指数分级': index_level,  # 新增：跟踪指数分级
            '跟踪ETF数量': stats['跟踪ETF数量'],
            '商务品数量': stats['商务品数量'],
            '商务品合计规模': stats['商务品合计规模'],
            '合计规模': stats['合计规模'],
            '一级分类': group['一级分类'].iloc[0] if '一级分类' in group.columns else None,
            '二级分类': group['二级分类'].iloc[0] if '二级分类' in group.columns else None,
            '三级分类': group['三级分类'].iloc[0] if '三级分类' in group.columns else None,
            '综合费率最低的基金': lowest_fee_fund['证券简称'],
            '综合费率最低的基金代码': lowest_fee_fund['证券代码_处理'],
            '综合费率最低的基金公司': simplify_fund_company_name(lowest_fee_fund['基金管理人']),
            '日均交易量最大的基金': highest_volume_fund['证券简称'],
            '日均交易量最大的基金代码': highest_volume_fund['证券代码_处理'],
            '日均交易量最大的基金公司': simplify_fund_company_name(highest_volume_fund['基金管理人']),
            '规模合计最大的基金': largest_scale_fund['证券简称'],
            '规模合计最大的基金代码': largest_scale_fund['证券代码_处理'],
            '规模合计最大的基金公司': simplify_fund_company_name(largest_scale_fund['基金管理人'])
        })
    except Exception as e:
        print(f"处理指数 {index_code} 时出错: {str(e)}")
        continue

# 创建DataFrame并处理
results_df = pd.DataFrame(results)
# 确保 '合计规模' 列存在
if '合计规模' not in results_df.columns:
    raise KeyError("'合计规模' 列未正确添加到 results_df 中")

results_df = calculate_category_totals(results_df)

# 列顺序调整
columns_order = [
    '跟踪指数代码', '跟踪指数名称', '跟踪指数分级', '跟踪ETF数量', '商务品数量', '商务品合计规模', '合计规模', '一级分类', '二级分类', '二级分类合计',
    '三级分类', '三级分类合计', '综合费率最低的基金', '综合费率最低的基金代码',
    '综合费率最低的基金公司', '日均交易量最大的基金', '日均交易量最大的基金代码',
    '日均交易量最大的基金公司', '规模合计最大的基金', '规模合计最大的基金代码',
    '规模合计最大的基金公司'
]
results_df = results_df[columns_order]

# 最终排序
sorted_results = custom_sort(results_df)


# 在 apply_excel_styles 函数中处理 '合计规模' 列
def apply_excel_styles(worksheet, df):
    # 数值格式设置
    num_cols = ['商务品合计规模', '合计规模', '二级分类合计', '三级分类合计']  # 添加 '合计规模'
    for col in num_cols:
        col_idx = df.columns.get_loc(col) + 1
        for row in range(2, worksheet.max_row + 1):
            worksheet.cell(row=row, column=col_idx).number_format = '0.0'

    # 获取各关键列索引
    scale_col = df.columns.get_loc('商务品合计规模') + 1  # 商务品合计规模
    total_scale_col = df.columns.get_loc('合计规模') + 1  # 合计规模
    count_col = df.columns.get_loc('跟踪ETF数量') + 1
    code_col = df.columns.get_loc('跟踪指数代码') + 1
    code_columns = [
        df.columns.get_loc('综合费率最低的基金代码') + 1,
        df.columns.get_loc('日均交易量最大的基金代码') + 1,
        df.columns.get_loc('规模合计最大的基金代码') + 1
    ]

    for row in range(2, worksheet.max_row + 1):
        # 需求2：商务品合计规模颜色
        scale_value = worksheet.cell(row=row, column=scale_col).value or 0
        scale_font = Font(bold=True)
        if scale_value >= 1000:
            scale_font.color = '8B0000'  # 深红
        elif 500 <= scale_value < 1000:
            scale_font.color = 'FFA500'  # 橙
        elif 200 <= scale_value < 500:
            scale_font.color = 'FFD700'  # 深黄
        elif 100 <= scale_value < 200:
            scale_font.color = '008000'  # 绿
        elif 50 <= scale_value < 100:
            scale_font.color = '00FFFF'  # 青
        elif 20 <= scale_value < 50:
            scale_font.color = '0000FF'  # 蓝
        elif 10 <= scale_value < 20:
            scale_font.color = '800080'  # 紫
        elif 5 <= scale_value < 10:
            scale_font.color = '000000'  # 黑
        elif 2 <= scale_value < 5:
            scale_font.color = 'A9A9A9'  # 深灰
        else:
            scale_font.color = '808080'  # 灰
        worksheet.cell(row=row, column=scale_col).font = scale_font

        # 需求3：ETF数量颜色
        count_value = worksheet.cell(row=row, column=count_col).value or 0
        count_font = Font(bold=True)
        if count_value >= 20:
            count_font.color = '8B0000'  # 深红
        elif 16 <= count_value < 20:
            count_font.color = 'FFA500'  # 橙
        elif 10 <= count_value < 16:
            count_font.color = 'FFD700'  # 深黄
        elif 7 <= count_value < 10:
            count_font.color = '008000'  # 绿
        elif 5 <= count_value < 7:
            count_font.color = '00FFFF'  # 青
        elif 3 <= count_value < 5:
            count_font.color = '0000FF'  # 蓝
        elif count_value == 2:
            count_font.color = '800080'  # 紫
        else:
            count_font.color = '000000'  # 黑
        worksheet.cell(row=row, column=count_col).font = count_font

        # 需求4：跟踪指数代码颜色
        codes = [
            str(worksheet.cell(row=row, column=code_columns[0]).value).strip(),
            str(worksheet.cell(row=row, column=code_columns[1]).value).strip(),
            str(worksheet.cell(row=row, column=code_columns[2]).value).strip()
        ]
        business = [code in business_etf_codes for code in codes]

        code_font = Font(bold=True)
        if all(business):
            code_font.color = '8B0000'  # 深红
        elif business[0] and business[1]:
            code_font.color = 'FFA500'  # 橙
        elif business[1] and business[2]:
            code_font.color = 'FFD700'  # 深黄
        elif business[2] and business[0]:
            code_font.color = '008000'  # 绿
        elif business[1]:
            code_font.color = '00FFFF'  # 青
        elif business[0]:
            code_font.color = '0000FF'  # 蓝
        elif business[2]:
            code_font.color = '800080'  # 紫
        else:
            code_font.color = '000000'  # 黑
        worksheet.cell(row=row, column=code_col).font = code_font

    # 新增：商务品代码标红功能
    business_code_cols = [
        df.columns.get_loc('综合费率最低的基金代码') + 1,
        df.columns.get_loc('日均交易量最大的基金代码') + 1,
        df.columns.get_loc('规模合计最大的基金代码') + 1
    ]
    red_bold = Font(color="FF0000", bold=True)  # 标准红色加粗

    for row in range(2, worksheet.max_row + 1):
        # 原有其他样式设置...

        # 新增商务品标红逻辑
        for col in code_columns:
            cell = worksheet.cell(row=row, column=col)
            code_value = str(cell.value).strip()
            if code_value in business_etf_codes:
                cell.font = red_bold


# 保存结果
with pd.ExcelWriter('THE_BEST_ETF_FINAL.xlsx', engine='openpyxl') as writer:
    sorted_results.to_excel(writer, index=False, sheet_name='优选ETF')
    apply_excel_styles(writer.sheets['优选ETF'], sorted_results)

print("处理完成，结果已保存至 THE_BEST_ETF_FINAL.xlsx")

# 规模分层与分级统计
# 规模分层（从小到大排列）
scale_bins = [0, 5, 10, 20, 50, 100, 300, 1000, float('inf')]
scale_labels = ['<5亿', '5-10亿', '10-20亿', '20-50亿', '50-100亿', '100-300亿', '300-1000亿', '>=1000亿']

# 添加规模分层列
results_df['指数规模分层'] = pd.cut(
    results_df['合计规模'],
    bins=scale_bins,
    labels=scale_labels,
    right=False
)

# 确保从一级到九级的所有列都存在
all_levels = ['一级', '二级', '三级', '四级', '五级', '六级', '七级', '八级', '九级']
for level in all_levels:
    if level not in results_df['跟踪指数分级'].unique():
        # 如果某个级别不存在，则添加一个空行
        results_df = results_df.append({'跟踪指数分级': level}, ignore_index=True)

# 统计各规模层级下，一到九级跟踪指数分级的个数
scale_quality_table = pd.crosstab(
    results_df['指数规模分层'],
    results_df['跟踪指数分级'],
    margins=True,  # 添加行和列的合计
    margins_name='合计'  # 合计列的名称
)

# 确保所有分级列都存在，即使某些分级的统计结果为 0
for level in all_levels:
    if level not in scale_quality_table.columns:
        scale_quality_table[level] = 0

# 按规模分层从小到大排列
scale_quality_table = scale_quality_table.reindex(scale_labels + ['合计'])

# 按跟踪指数分级从一级到九级排列
scale_quality_table = scale_quality_table[all_levels + ['合计']]

# 保存结果到 Sheet2
with pd.ExcelWriter('THE_BEST_ETF_FINAL.xlsx', engine='openpyxl', mode='a') as writer:
    scale_quality_table.to_excel(writer, sheet_name='Tracking_Index_Scale&Quality')
    
    # 获取工作表对象
    worksheet = writer.sheets['Tracking_Index_Scale&Quality']
    
    # 添加跟踪指数分级逻辑说明
    row_offset = scale_quality_table.shape[0] + 3  # 在统计表下方留出空行
    
    # 添加标题
    worksheet.cell(row=row_offset, column=1, value="跟踪指数分级逻辑说明")
    worksheet.cell(row=row_offset, column=1).font = Font(bold=True, size=12)
    
    # 添加分级逻辑说明
    level_descriptions = [
        ["一级", "最高交易量基金、最低费率基金、最大规模基金均为商务品"],
        ["二级", "最高交易量基金、最低费率基金为商务品，最大规模基金非商务品"],
        ["三级", "最高交易量基金、最大规模基金为商务品，最低费率基金非商务品"],
        ["四级", "最低费率基金、最大规模基金为商务品，最高交易量基金非商务品"],
        ["五级", "仅最高交易量基金为商务品"],
        ["六级", "仅最大规模基金为商务品"],
        ["七级", "仅最低费率基金为商务品"],
        ["八级", "有商务品，但不是最高交易量/最低费率/最大规模基金"],
        ["九级", "无商务品"]
    ]
    
    # 写入分级逻辑
    for i, (level, description) in enumerate(level_descriptions):
        worksheet.cell(row=row_offset + 2 + i, column=1, value=level)
        worksheet.cell(row=row_offset + 2 + i, column=2, value=description)
        worksheet.cell(row=row_offset + 2 + i, column=1).font = Font(bold=True)

    # 添加额外说明
    extra_note_row = row_offset + 2 + len(level_descriptions) + 1
    worksheet.cell(row=extra_note_row, column=1, value="注意事项")
    worksheet.cell(row=extra_note_row, column=1).font = Font(bold=True)
    
    extra_note = "有些指数只有一个跟踪的ETF产品，如果该ETF是我司商务品，则会体现为一级。但跟踪少往往代表指数规模较小，这样的一级含金量不算高。比如小于5亿规模的指数往往是这种情况。"
    worksheet.cell(row=extra_note_row, column=2, value=extra_note)
    worksheet.cell(row=extra_note_row, column=2).font = Font(italic=True)

print("规模与分级统计表已保存至 Sheet2: Tracking_Index_Scale&Quality，并添加了分级逻辑说明")